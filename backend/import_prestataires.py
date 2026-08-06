import os
import sys
import django

# Setup Django environment
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from crm.models import Company, PrestataireEvaluation

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Evaluation des prestataires.xlsx')

def clean_val(v):
    if v is None:
        return ''
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s

def clean_num(v):
    if v is None or v == '' or v == 'None':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def run_import():
    if not os.path.exists(EXCEL_PATH):
        print(f"File not found at {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print("Loaded workbook sheets:", wb.sheetnames)

    # 1. Import Companies from "Liste de Prestataires " and "Copie de Liste de Prestataires "
    company_map = {}

    for sheet_name in ['Liste de Prestataires ', 'Copie de Liste de Prestataires ']:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        start_row = 10 if 'Copie' in sheet_name else 7
        
        for row in list(ws.iter_rows(values_only=True))[start_row-1:]:
            if not row or len(row) < 3:
                continue
            
            if 'Copie' in sheet_name:
                # Cols: ID(0), Nom(1), Patente(2), Domaine(3), Interlocuteur(4), Fonction(5), Contact(6), Convention(7), Deb(8), Fin(9), Notes(10)
                name = clean_val(row[1])
                patente = clean_val(row[2])
                domaine = clean_val(row[3])
                interlocuteur = clean_val(row[4])
                fonction = clean_val(row[5])
                contact = clean_val(row[6])
                convention = clean_val(row[7])
                debut = clean_val(row[8])
                fin = clean_val(row[9])
                notes = clean_val(row[10])
            else:
                # Cols: ID(0), Nom(1), Patente(2), Domaine(3), Interlocuteur(4), Fonction(5), Contact(6), CV(7), Convention(8), Deb(9), Fin(10)
                name = clean_val(row[1])
                patente = clean_val(row[2])
                domaine = clean_val(row[3])
                interlocuteur = clean_val(row[4])
                fonction = clean_val(row[5])
                contact = clean_val(row[6])
                convention = clean_val(row[8])
                debut = clean_val(row[9])
                fin = clean_val(row[10])
                notes = ''

            if not name or name.lower().startswith('nom'):
                continue

            email = contact if '@' in contact else ''
            phone = contact if '@' not in contact and len(contact) > 3 else ''

            company, created = Company.objects.get_or_create(
                name=name,
                defaults={
                    'status': 'Prestataire',
                    'service_provided': domaine or '',
                    'activity_domain': domaine,
                    'patente': patente if patente else 'Oui',
                    'interlocuteur_principal': interlocuteur,
                    'fonction_interlocuteur': fonction,
                    'address': '',
                    'convention_contrat': convention if convention else 'Non',
                    'email': email,
                    'phone': phone,
                }
            )

            # Update existing if needed
            company.status = 'Prestataire'
            if patente: company.patente = patente
            company.service_provided = domaine or ''
            if interlocuteur: company.interlocuteur_principal = interlocuteur
            if fonction: company.fonction_interlocuteur = fonction
            if convention: company.convention_contrat = convention
            if email: company.email = email
            if phone: company.phone = phone
            company.save()

            company_map[name] = company
            print(f"{'Created' if created else 'Updated'} Prestataire: {name}")

    # 2. Import Evaluations (2024, 2025, 2026)
    for year in [2024, 2025, 2026]:
        sheet_name = f'Evaluation {year}'
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        # Find header row
        header_idx = -1
        for i, row in enumerate(rows):
            if row and any('Nom' in str(cell) for cell in row if cell):
                header_idx = i
                break
        
        if header_idx == -1:
            header_idx = 5

        for row in rows[header_idx+1:]:
            if not row or len(row) < 6:
                continue
            
            name = clean_val(row[0])
            if not name or name.lower().startswith('nom'):
                continue

            # Ensure company exists
            company = company_map.get(name)
            if not company:
                company, _ = Company.objects.get_or_create(
                    name=name,
                    defaults={
                        'status': 'Prestataire',
                        'patente': clean_val(row[1]) or 'Oui',
                        'service_provided': clean_val(row[2]) or 'Autre',
                        'interlocuteur_principal': clean_val(row[3]),
                        'fonction_interlocuteur': clean_val(row[4]),
                    }
                )
                company_map[name] = company

            prix = clean_num(row[6]) if len(row) > 6 else None
            qualite = clean_num(row[7]) if len(row) > 7 else None
            prestation = clean_num(row[8]) if len(row) > 8 else None
            delais = clean_num(row[9]) if len(row) > 9 else None
            notes = clean_val(row[11]) if len(row) > 11 else ''

            eval_obj, created = PrestataireEvaluation.objects.get_or_create(
                company=company,
                year=year,
                defaults={
                    'prix': prix,
                    'qualite': qualite,
                    'prestation_service': prestation,
                    'respect_delais': delais,
                    'notes': notes,
                }
            )

            if not created:
                eval_obj.prix = prix
                eval_obj.qualite = qualite
                eval_obj.prestation_service = prestation
                eval_obj.respect_delais = delais
                if notes: eval_obj.notes = notes
                eval_obj.save()

            print(f"Saved evaluation {year} for {name}: score={eval_obj.score_percent}%, decision={eval_obj.decision}")

if __name__ == '__main__':
    run_import()
